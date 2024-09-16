from openpyxl import Workbook
from rest_framework import status
from django.http import HttpResponse
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

from apps.diet.models import Diet
from apps.user.models import Athlete
from apps.core.permissions import IsAthleteUser
from apps.diet.serializers import DietSerializer
from apps.daily_records.models import DailyRecords
from apps.core.mixins import AthleteNutritionistPermissionMixin


class DietViewSet(AthleteNutritionistPermissionMixin):
    serializer_class = DietSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset_model_class(self):
        return Diet

    def get_related_model_id(self):
        return self.request.data.get('athlete')

    def get_related_model_class(self):
        return Athlete

    @action(detail=False, methods=['get'], url_path='diet-count', permission_classes=[IsAuthenticated, IsAthleteUser])
    def diet_dates(self, request):
        diets = Diet.objects.all().values('id', 'initial_date', 'final_date')
        return Response(diets)


    def destroy(self, request, *args, **kwargs):
        diet = self.get_object()
        daily_records = DailyRecords.objects.filter(meal__diet=diet)

        if daily_records.exists():
            return Response(
                {"detail": "Cannot delete this diet because the athlete has already started executing it and has daily records associated with its meals."},
                status=status.HTTP_400_BAD_REQUEST
            )

        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['get'], url_path='export', permission_classes=[IsAuthenticated])
    def export_diet(self, request, pk=None):
        diet = self.get_object()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=diet_{diet.id}.xlsx'

        # Criação da planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Diet"

        # Remover as linhas de grade
        ws.sheet_view.showGridLines = False

        # Definindo estilos de preenchimento de cores
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

        # Estilos de borda e alinhamento
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Adicionando meta, observações e datas
        ws['A1'] = f"Meta: {diet.goal}"
        ws['A2'] = f"Observações: {diet.observations}"
        ws['A3'] = f"Data Inicial: {diet.initial_date.strftime('%Y-%m-%d')}"
        ws['A4'] = f"Data Final: {diet.final_date.strftime('%Y-%m-%d')}"

        # Mesclando as células da coluna A e B
        ws.merge_cells('A1:I1')
        ws.merge_cells('A2:I2')
        ws.merge_cells('A3:I3')
        ws.merge_cells('A4:I4')

        # Estilos das células de Meta e Observações
        for cell in ['A1', 'A2', 'A3', 'A4']:
            ws[cell].alignment = center_alignment
            ws[cell].border = thin_border
            ws[cell].fill = orange_fill

        # Espaçamento para cabeçalho da tabela de alimentos
        ws.append([""])

        # Adicionando cabeçalhos da tabela
        headers = ["Refeição", "Horário", "Alimento", "kcal", "PTN", "CHO", "LIP", "Quantidade", "Fibras (g)"]
        ws.append(headers)

        # Estilo dos cabeçalhos
        header_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            header_cell = ws.cell(row=6, column=col)
            header_cell.font = header_font
            header_cell.alignment = center_alignment
            header_cell.border = thin_border
            header_cell.fill = yellow_fill

        # Definindo largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40

        row_num = 7
        for meal in diet.meals.all():
            first_row = row_num
            foods = meal.foods  # Acessando os alimentos da refeição (JSONField)

            # Verificar se foods é uma lista de dicionários
            if isinstance(foods, list) and foods:
                for food in foods:
                    ws.append([
                        meal.name,
                        meal.time,
                        food.get('food_description', ''),
                        food.get('energy_kcal', 0),
                        food.get('protein', 0),
                        food.get('carbohydrates', 0),
                        food.get('lipids', 0),
                        food.get('quantity', 0),
                        food.get('dietary_fiber', 0)
                    ])
                    row_num += 1
            else:
                # Caso não tenha alimentos na refeição
                ws.append([
                    meal.name,
                    meal.time,
                    "Sem alimentos",
                    '', '', '', '', '', ''
                ])
                row_num += 1

            # Verificar se a refeição tem mais de um alimento antes de mesclar
            if row_num - first_row > 1:
                ws.merge_cells(start_row=first_row, start_column=1, end_row=row_num - 1, end_column=1)
                ws.merge_cells(start_row=first_row, start_column=2, end_row=row_num - 1, end_column=2)

            # Aplicando bordas e alinhamento às células
            for col in range(1, len(headers) + 1):
                for row in range(first_row, row_num):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = center_alignment

                # Aplicando cor de fundo às células de dados
                if meal.name.lower() == "café da manhã":
                    ws.cell(row=row_num - 1, column=1).fill = pink_fill
                else:
                    ws.cell(row=row_num - 1, column=1).fill = blue_fill

            # Adicionar informações do MealMacrosSheet
            meal_macros = meal.meal_macros_sheet
            ws.append(
                ["", "", "Macros da Refeição", meal_macros.kcal, meal_macros.ptn, meal_macros.cho, meal_macros.fat]
            )
            row_num += 1

            # Adicionar uma linha em branco após os macros da refeição
            ws.append([""])
            row_num += 1

        # Adicionar espaço após as refeições
        ws.append([""])
        row_num += 1

        # Adicionar informações do DietMacrosSheet
        diet_macros = diet.diet_macros_sheet
        ws.append(["", "", "Macros da Dieta", diet_macros.kcal, diet_macros.ptn, diet_macros.cho, diet_macros.fat])
        ws.append(["", "", "Proporções", f"CHO: {diet_macros.cho_proportion}%", f"PTN: {diet_macros.ptn_proportion}%",
                   f"FAT: {diet_macros.fat_proportion}%"])

        # Aplicar cores às fontes dos valores de macro
        font_protein = Font(color="FF69B4")  # Rosa
        font_carbs = Font(color="00FF00")  # Verde
        font_fat = Font(color="FFFF00")  # Amarelo

        for row in ws.iter_rows(min_row=7, min_col=4, max_col=9):
            for cell in row:
                if 'CHO' in ws.cell(row=6, column=cell.column).value:
                    cell.font = font_carbs
                elif 'PTN' in ws.cell(row=6, column=cell.column).value:
                    cell.font = font_protein
                elif 'LIP' in ws.cell(row=6, column=cell.column).value:
                    cell.font = font_fat

        # Salvando o arquivo
        wb.save(response)
        return response